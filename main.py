#!/usr/bin/env python3
"""
Конвертирует все .xlsx файлы в CSV с корректной обработкой объединённых ячеек.
Значения из объединённого диапазона дублируются во все входящие в него ячейки.
"""

import argparse
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries


def fill_merged_cells(ws):
    """
    Для рабочего листа openpyxl «разливает» значения из объединённых ячеек
    по всем ячейкам диапазона. Возвращает заполненный лист (изменяется на месте).
    """
    merges = list(ws.merged_cells.ranges)  # фиксируем список, т.к. при изменении может меняться
    for merge_range in merges:
        min_col, min_row, max_col, max_row = (
            int(v) for v in range_boundaries(str(merge_range))
        )
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merge_range))  # убираем объединение
        # Заполняем все ячейки диапазона значением из верхнего левого угла
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value


def sheet_to_dataframe(ws, header_row=1):
    """
    Преобразует лист openpyxl в DataFrame после «разлива» объединённых ячеек.
    Первая строка (header_row по умолчанию 1) считается заголовком.
    """
    # Определяем реальный диапазон данных (игнорируем пустые строки/столбцы за пределами)
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row is None or max_col is None:
        return pd.DataFrame()

    # Собираем данные в список списков
    data = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        data.append(list(row))

    if not data:
        return pd.DataFrame()

    # Определяем строку заголовков
    header_idx = header_row - 1  # перевод в 0‑индексацию
    if header_idx >= len(data):
        return pd.DataFrame(data)  # заголовка нет

    columns = data[header_idx]
    # Всё, что ниже заголовка — данные
    data_rows = data[header_idx + 1 :]

    # Приводим колонки к строкам (если есть None → пустая строка)
    str_columns = [str(c) if c is not None else '' for c in columns]

    df = pd.DataFrame(data_rows, columns=str_columns)
    return df


def convert_xlsx_to_csv(directory: str) -> None:
    root = Path(directory)
    if not root.is_dir():
        print(f"Ошибка: '{directory}' не является директорией.")
        return

    xlsx_files = list(root.glob("*.xlsx"))
    if not xlsx_files:
        print(f"В каталоге '{root}' нет файлов .xlsx.")
        return

    for xlsx_path in xlsx_files:
        print(f"Обрабатываю: {xlsx_path.name}")
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                fill_merged_cells(ws)          # исправляем объединения
                df = sheet_to_dataframe(ws)    # первая строка → заголовок
                if df.empty:
                    print(f"  Лист '{sheet_name}' пуст, пропускаю.")
                    continue

                # Имя csv-файла
                stem = xlsx_path.stem
                safe_sheet = sheet_name.replace("/", "_").replace("\\", "_")
                csv_name = f"{stem}_{safe_sheet}.csv"
                csv_path = xlsx_path.parent / csv_name

                df.to_csv(csv_path, index=False, encoding="utf-8-sig")
                print(f"  Сохранён: {csv_path.name} ({len(df)} строк)")
        except Exception as e:
            print(f"  Ошибка при обработке {xlsx_path.name}: {e}")


def main():
    convert_xlsx_to_csv("macro")


if __name__ == "__main__":
    main()